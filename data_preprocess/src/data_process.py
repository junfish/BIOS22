from openpyxl import load_workbook, Workbook
import pickle
# ADNI1GO2_MRI_FDG_AV45.xlsx and demographic_information_combined-v2.xlsx are two original data files

AV45_workbook = load_workbook(filename = "../data/data_original/ADNI1GO2_MRI_FDG_AV45.xlsx")
AV45_sheet = AV45_workbook.active
AV45_rows = AV45_sheet.max_row
AV45_cols = AV45_sheet.max_column

DemoInfo_workbook = load_workbook(filename = "../data/data_original/demographic_information_combined-v2.xlsx")
DemoInfo_sheet = DemoInfo_workbook.active
DemoInfo_rows = DemoInfo_sheet.max_row
DemoInfo_cols = DemoInfo_sheet.max_column

Match_wb = Workbook()
ws = Match_wb.create_sheet(title = "Match_IID", index = 0)
match_row = 1
for index, attr in enumerate(['IID', 'date_diff', 'DXGrp', 'diagnosis', 'adas13', 'MMSE']):
    ws.cell(match_row, index + 1).value = attr

for row_idx1 in range(2, AV45_rows + 1):
    # prepare data in ADNI1GO2_MRI_FDG_AV45.xlsx
    # attributes in ADNI1GO2_MRI_FDG_AV45.xlsx: [IID, Sex, Age, Edu, APOEGrp, DXGrp, MRI_1_Date, AV45_1_Date, FDG_1_Date]
    date_list = []
    IID_str1 = AV45_sheet.cell(row = row_idx1, column = 1).value.replace('_', '') # Process IID from 023_S_0031 to 023S0031, for example.
    for col_idx in range(7, 10):
        if AV45_sheet.cell(row = row_idx1, column = col_idx).value:
            date_list.append(AV45_sheet.cell(row = row_idx1, column = col_idx).value)
    date2ordinal_list = [date.toordinal() for date in date_list] # Obtain three dates as an order number
    # if len(date_list) == 0:
    #     print("!!!!!")

    # prepare data in demographic_information_combined-v2.xlsx
    min_diff_abs = 9999999
    for row_idx2 in range(2, DemoInfo_rows + 1):
        # prepare data in ADNI1GO2_MRI_FDG_AV45.xlsx
        # attributes in ADNI1GO2_MRI_FDG_AV45.xlsx: [filename, session_id, examination_date, age, ApoeGrp, Edu, Sex, current_study, diagnosis, adas13, MMSE]
        IID_str2 = DemoInfo_sheet.cell(row = row_idx2, column = 1).value.replace('sub-ADNI', '') # Process IID from sub-ADNI002S0295 to 002S0295, for example.
        if IID_str1 == IID_str2: # Match IID & filename to connect DxGrp (AD stage), adas13 and MMSE
            if DemoInfo_sheet.cell(row = row_idx2, column = 3).value:
                date2ordinal = DemoInfo_sheet.cell(row = row_idx2, column = 3).value.toordinal() # Obtain date as an order number
            else:
                continue

            diff_abs = min([abs(date2ordinal - _) for _ in date2ordinal_list])
            # if any value of DXGrp in ADNI1GO2_MRI_FDG_AV45.xlsx and adas13 & MMSE in demographic_information_combined-v2.xlsx is none, this matching is useless.
            if_any_none = (AV45_sheet.cell(row = row_idx1, column = 6).value == None) or \
                          (DemoInfo_sheet.cell(row = row_idx2, column = 10).value == None) or \
                          (DemoInfo_sheet.cell(row = row_idx2, column = 11).value == None)
            if (diff_abs < min_diff_abs) and not(if_any_none):
                min_diff_abs = diff_abs
                new_list = [IID_str1,
                            min_diff_abs,
                            AV45_sheet.cell(row=row_idx1, column=6).value,
                            DemoInfo_sheet.cell(row=row_idx2, column=9).value,
                            DemoInfo_sheet.cell(row=row_idx2, column=10).value,
                            DemoInfo_sheet.cell(row=row_idx2, column=11).value]
    if new_list:
        match_row += 1
        ws.cell(match_row, 1).value = new_list[0]
        ws.cell(match_row, 2).value = new_list[1]
        ws.cell(match_row, 3).value = new_list[2]
        ws.cell(match_row, 4).value = new_list[3]
        ws.cell(match_row, 5).value = new_list[4]
        ws.cell(match_row, 6).value = new_list[5]


Match_wb.save("../data/data_after_process/Match_date_diff_IID.xlsx")
# with open("../data/Match_IID.txt", "wb") as fp:   #Pickling
#     pickle.dump(Match_list, fp)
print("finished! This is the first step to process the data: match the data via the IID and filename (i.e., patient/participant)")
