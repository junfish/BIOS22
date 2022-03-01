from openpyxl import load_workbook, Workbook
import scipy.io as sio

Final_wb = Workbook()
ws = Final_wb.create_sheet(title = "Final_sbjIID_scores", index = 0)
match_row = 1
for index, attr in enumerate(['IID', 'AD_label', 'adas13', 'MMSE']):
    ws.cell(match_row, index + 1).value = attr

match_wb = load_workbook("../data/data_after_process/Match2nd_date_diff_trans_IID.xlsx")
match_ws = match_wb.active
match_rows = match_ws.max_row
match_cols = match_ws.max_column

for row_num in range(2, match_rows + 1):
    if match_ws.cell(row_num, 2).value < 15 and match_ws.cell(row_num, 3).value == match_ws.cell(row_num, 4).value: # define a tolerance of date difference or the tolerance of data error (AD_label & DXGrp)
    # and match_ws.cell(row_num, 3).value == match_ws.cell(row_num, 4).value:
        match_row += 1

        ws.cell(match_row, 1).value = str(match_ws.cell(row_num, 1).value)
        ws.cell(match_row, 2).value = match_ws.cell(row_num, 3).value
        ws.cell(match_row, 3).value = match_ws.cell(row_num, 6).value
        ws.cell(match_row, 4).value = match_ws.cell(row_num, 7).value
print("The final dataset contains " + str(match_row - 1) + " participants")

Final_wb.save("../data/data_after_process/Final_dataset_14_date_diff_no_error.xlsx")

# < 1 --> 388 participants
# < 1 and no error --> 365 participants

# < 8 --> 595 participants
# < 8 and no error --> 566 participants

# < 15 --> 680 participants
# < 15 and no error --> 649 participants

# < 21 --> 709 participants
# < 21 and no error --> 674 participants

# < 31 --> 725 participants
# < 31 and no error --> 688 participants

# < 61 --> 731 participants
# < 61 and no error --> 694 participants   %% unique: 692

# < 81 --> 735 participants
# < 81 and no error --> 697 participants