from openpyxl import load_workbook, Workbook
import scipy.io as sio
# Match ../data/data_after_process/Match_date_diff_IID.xlsx with ADNI.mat via the IID and sbjID

labels = sio.loadmat("../data/data_original/labels.mat")["labels"] # AD stage stored in ADNI.mat
sbjID = sio.loadmat("../data/data_original/sbjID.mat")["sbjID"] # sbjID stored in ADNI.mat
sbjID_list = []
for item in sbjID:
    sbjID_list.append(str(item[0]).split('\'')[1][1:].replace('_', '')) # Process IID from 1023_S_0031 to 023S0031

sbj_label_dict = dict(zip(sbjID_list, labels)) # Zip the sbjID and its AD stage label

match_wb = load_workbook("../data/data_after_process/Match_date_diff_IID.xlsx") # final save directory in the script data_process.py
match_ws = match_wb.active
match_rows = match_ws.max_row
match_cols = match_ws.max_column
num = 0

Match2nd_wb = Workbook()
ws = Match2nd_wb.create_sheet(title = "Match_sbjIID", index = 0)
match_row = 1
for index, attr in enumerate(['IID', 'date_diff', 'AD_label', 'DXGrp', 'diagnosis', 'adas13', 'MMSE']): # Add a new column 'AD_label' to compare with transformed DXGrp
    ws.cell(match_row, index + 1).value = attr

for row_num in range(2, match_rows + 1):
    if match_ws.cell(row_num, 1).value in sbjID_list:
        match_row += 1
        ws.cell(match_row, 1).value = match_ws.cell(row_num, 1).value
        ws.cell(match_row, 2).value = match_ws.cell(row_num, 2).value
        ws.cell(match_row, 3).value = sbj_label_dict[match_ws.cell(row_num, 1).value][0]
        if 1 < match_ws.cell(row_num, 3).value < 5:
            ws.cell(match_row, 4).value = match_ws.cell(row_num, 3).value + 1 # 2 -> 3; 3 ->4; 4 -> 5
        elif match_ws.cell(row_num, 3).value == 5: # 5 -> 2
            ws.cell(match_row, 4).value = 2
        else:
            ws.cell(match_row, 4).value = match_ws.cell(row_num, 3).value # 1 -> 1

        ws.cell(match_row, 5).value = match_ws.cell(row_num, 4).value
        ws.cell(match_row, 6).value = match_ws.cell(row_num, 5).value
        ws.cell(match_row, 7).value = match_ws.cell(row_num, 6).value

Match2nd_wb.save("../data/data_after_process/Match2nd_date_diff_trans_IID.xlsx")
print("finished! This is the second step to precess the data: match the data via the sbjID and IID, and transform the DXGrp to be consistent with ADNI.mat.")